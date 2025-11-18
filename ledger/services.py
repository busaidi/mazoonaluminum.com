# ledger/services.py
from datetime import timedelta
from decimal import Decimal

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook

from ledger.models import Account, JournalLine, JournalEntry, Journal, FiscalYear, LedgerSettings


def build_lines_from_formset(line_formset):
    """
    يحوّل الفورم سِت لليست أسطر جاهزة للإنشاء + يحسب الإجماليات.

    يعتمد على:
    - clean() في JournalLineForm للتحقق الأساسي (القيم السالبة، مدين/دائن معاً، إلخ).
    - يتعامل مع DELETE + الأسطر الفارغة.

    يعيد:
    - lines: قائمة dict فيها بيانات كل سطر.
    - total_debit: مجموع المدين.
    - total_credit: مجموع الدائن.

    يرفع ValidationError في حال:
    - لا يوجد أي سطر صالح.
    """

    lines = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    has_any_line = False

    for order, form in enumerate(line_formset.forms):
        if form.cleaned_data.get("DELETE"):
            continue

        account = form.cleaned_data.get("account")
        description = form.cleaned_data.get("description") or ""
        debit = form.cleaned_data.get("debit") or Decimal("0")
        credit = form.cleaned_data.get("credit") or Decimal("0")

        # سطر فارغ بالكامل → تجاهل
        if not account and not description and debit == 0 and credit == 0:
            continue

        has_any_line = True

        lines.append(
            {
                "account": account,
                "description": description,
                "debit": debit,
                "credit": credit,
                "order": order,
            }
        )

        total_debit += debit
        total_credit += credit

    if not has_any_line:
        raise ValidationError(_("لا يوجد أي سطر صالح في القيد."))

    return lines, total_debit, total_credit


def ensure_default_chart_of_accounts():
    """
    ينشئ شجرة حسابات افتراضية أساسية إذا لم يكن هناك أي حساب في النظام.

    يعيد عدد الحسابات التي تم إنشاؤها.
    """

    if Account.objects.exists():
        # يوجد حسابات مسبقاً → لا نعمل شيء
        return 0

    accounts_data = [
        # code, name, type, allow_settlement
        ("1000", _("الصندوق"), Account.Type.ASSET, True),
        ("1010", _("البنك"), Account.Type.ASSET, True),
        ("1100", _("العملاء"), Account.Type.ASSET, True),
        ("1200", _("المخزون"), Account.Type.ASSET, False),

        ("2000", _("الموردون"), Account.Type.LIABILITY, True),

        ("3000", _("رأس المال"), Account.Type.EQUITY, False),

        ("4000", _("مبيعات"), Account.Type.REVENUE, False),

        ("5000", _("مشتريات"), Account.Type.EXPENSE, False),
        ("5100", _("مصروفات عمومية وإدارية"), Account.Type.EXPENSE, False),
    ]

    created_count = 0
    for code, name, acc_type, allow_settlement in accounts_data:
        Account.objects.create(
            code=code,
            name=name,
            type=acc_type,
            allow_settlement=allow_settlement,
        )
        created_count += 1

    return created_count



def _parse_bool(value):
    """
    Helper to parse boolean-ish values from Excel.
    Accepted truthy values: 1, true, yes, y (case-insensitive).
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if not s:
        return None
    return s in {"1", "true", "yes", "y", "نعم", "صح"}


def import_chart_of_accounts_from_excel(
    file_obj,
    *,
    replace_existing: bool = False,
    fiscal_year: FiscalYear | None = None,
):
    """
    Import chart of accounts from an Excel file (xlsx).

    Expected columns:
      - code, name, type (required)
      - parent_code, allow_settlement, is_active (optional)
      - opening_debit, opening_credit (optional; used for opening balance)

    Behavior:
      - If account code exists → update.
      - Else → create.
      - If replace_existing=True → accounts not in file: is_active=False.
      - If file contains opening balances:
        * If fiscal_year is None → ValidationError (لازم تختار سنة).
        * Else → create ONE opening JournalEntry in that fiscal year.
    """

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    # --- header mapping ---
    header_map = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue
        name = str(raw).strip().lower()
        if not name:
            continue
        header_map[name] = col

    required = ["code", "name", "type"]
    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValidationError(
            _("Missing required columns in Excel: %(cols)s")
            % {"cols": ", ".join(missing)}
        )

    rows = []
    seen_codes: set[str] = set()
    errors: list[str] = []

    type_map = {
        "asset": Account.Type.ASSET,
        "liability": Account.Type.LIABILITY,
        "equity": Account.Type.EQUITY,
        "revenue": Account.Type.REVENUE,
        "expense": Account.Type.EXPENSE,
    }

    # --- read rows ---
    for row_idx in range(2, ws.max_row + 1):
        code_val = ws.cell(row=row_idx, column=header_map["code"]).value
        if code_val is None or str(code_val).strip() == "":
            continue

        code = str(code_val).strip()
        name = str(ws.cell(row=row_idx, column=header_map["name"]).value or "").strip()
        type_raw = ws.cell(row=row_idx, column=header_map["type"]).value

        if not name:
            errors.append(f"Row {row_idx}: empty name for code '{code}'")
            continue

        if not type_raw:
            errors.append(f"Row {row_idx}: empty type for code '{code}'")
            continue

        type_str = str(type_raw).strip().lower()
        if type_str not in type_map:
            errors.append(
                f"Row {row_idx}: invalid type '{type_str}' for code '{code}'. "
                "Expected one of: asset, liability, equity, revenue, expense."
            )
            continue

        parent_code = None
        if "parent_code" in header_map:
            parent_cell = ws.cell(row=row_idx, column=header_map["parent_code"]).value
            if parent_cell is not None:
                parent_code = str(parent_cell).strip() or None

        allow_settlement = None
        if "allow_settlement" in header_map:
            allow_cell = ws.cell(
                row=row_idx, column=header_map["allow_settlement"]
            ).value
            allow_settlement = _parse_bool(allow_cell)

        is_active = None
        if "is_active" in header_map:
            active_cell = ws.cell(row=row_idx, column=header_map["is_active"]).value
            is_active = _parse_bool(active_cell)

        # opening balances
        opening_debit = Decimal("0")
        opening_credit = Decimal("0")
        if "opening_debit" in header_map:
            val = ws.cell(row=row_idx, column=header_map["opening_debit"]).value
            if val not in (None, ""):
                opening_debit = Decimal(str(val))
        if "opening_credit" in header_map:
            val = ws.cell(row=row_idx, column=header_map["opening_credit"]).value
            if val not in (None, ""):
                opening_credit = Decimal(str(val))

        rows.append(
            {
                "row_idx": row_idx,
                "code": code,
                "name": name,
                "type": type_map[type_str],
                "parent_code": parent_code,
                "allow_settlement": allow_settlement,
                "is_active": is_active,
                "opening_debit": opening_debit,
                "opening_credit": opening_credit,
            }
        )
        seen_codes.add(code)

    created = 0
    updated = 0
    deactivated = 0

    with transaction.atomic():
        # --- create/update accounts ---
        accounts_by_code = {a.code: a for a in Account.objects.all()}

        for row in rows:
            code = row["code"]
            name = row["name"]
            acc_type = row["type"]
            allow_settlement = row["allow_settlement"]
            is_active = row["is_active"]

            defaults = {
                "name": name,
                "type": acc_type,
            }
            if allow_settlement is not None:
                defaults["allow_settlement"] = allow_settlement
            if is_active is not None:
                defaults["is_active"] = is_active

            if code in accounts_by_code:
                acc = accounts_by_code[code]
                for field, value in defaults.items():
                    setattr(acc, field, value)
                acc.parent = None
                acc.save()
                updated += 1
            else:
                acc = Account.objects.create(
                    code=code,
                    parent=None,
                    **defaults,
                )
                accounts_by_code[code] = acc
                created += 1

        # --- set parents ---
        for row in rows:
            code = row["code"]
            parent_code = row["parent_code"]
            if not parent_code:
                continue

            acc = accounts_by_code.get(code)
            parent = accounts_by_code.get(parent_code)
            if not parent:
                errors.append(
                    f"Row {row['row_idx']}: parent_code '{parent_code}' not found for account '{code}'."
                )
                continue

            if parent.code == acc.code:
                errors.append(
                    f"Row {row['row_idx']}: account '{code}' cannot be its own parent."
                )
                continue

            acc.parent = parent
            acc.save(update_fields=["parent"])

        # --- deactivate missing accounts if needed ---
        if replace_existing:
            qs = Account.objects.exclude(code__in=seen_codes)
            deactivated = qs.update(is_active=False)

        # --- opening balance logic ---
        # اجمع الأرصدة الافتتاحية
        opening_lines = []
        total_debit = Decimal("0")
        total_credit = Decimal("0")

        for row in rows:
            od = row["opening_debit"]
            oc = row["opening_credit"]
            if od == 0 and oc == 0:
                continue

            opening_lines.append((accounts_by_code[row["code"]], od, oc))
            total_debit += od
            total_credit += oc

        if opening_lines:
            # لو فيه رصيد افتتاحي في الملف ولا تم اختيار سنة مالية → خطأ منطقي
            if fiscal_year is None:
                raise ValidationError(
                    _(
                        "الملف يحتوي على أرصدة افتتاحية، "
                        "لكن لم يتم اختيار سنة مالية للرصد الافتتاحي."
                    )
                )

            if total_debit != total_credit:
                raise ValidationError(
                    _(
                        "الرصيد الافتتاحي غير متوازن: إجمالي المدين (%(debit)s) "
                        "لا يساوي إجمالي الدائن (%(credit)s)."
                    )
                    % {"debit": total_debit, "credit": total_credit}
                )

            # اختر دفتر الرصيد الافتتاحي من إعدادات الليدجر إن وجد،
            # وإذا مش مضبوط → استخدم الدفتر اليدوي الافتراضي،
            # وإذا حتى هذا مش مضبوط → ارجع للسلوك القديم get_default_for_manual_entry()
            settings_obj = LedgerSettings.get_solo()
            journal = (
                settings_obj.opening_balance_journal
                or settings_obj.default_manual_journal
                or Journal.objects.get_default_for_manual_entry()
            )

            # المرجع يبقى مرتبط بالسنة الجديدة (سنة الهدف)
            ref = f"OPENING-{fiscal_year.year}"

            # حاول نربط القيد بالسنة السابقة إن وجدت
            prev_fy = (
                FiscalYear.objects.filter(year=fiscal_year.year - 1)
                .order_by("-id")
                .first()
            )
            if prev_fy:
                entry_fiscal_year = prev_fy
                opening_date = prev_fy.end_date
            else:
                # لو ما فيه سنة سابقة، نخليها قبل بداية السنة بيوم وبدون سنة مالية
                entry_fiscal_year = None
                opening_date = fiscal_year.start_date - timedelta(days=1)

            # امسح أي قيد افتتاحي سابق بنفس المرجع (بغض النظر عن السنة المالية)
            JournalEntry.objects.filter(
                journal=journal,
                reference=ref,
            ).delete()

            entry = JournalEntry.objects.create(
                fiscal_year=entry_fiscal_year,
                journal=journal,
                date=opening_date,
                reference=ref,
                description=_("رصيد افتتاحي مستورد من إكسل"),
                posted=True,
                posted_at=timezone.now(),
            )

            order = 1
            for acc, od, oc in opening_lines:
                JournalLine.objects.create(
                    entry=entry,
                    account=acc,
                    description=_("رصيد افتتاحي"),
                    debit=od,
                    credit=oc,
                    order=order,
                )
                order += 1

    return {
        "created": created,
        "updated": updated,
        "deactivated": deactivated,
        "errors": errors,
    }

