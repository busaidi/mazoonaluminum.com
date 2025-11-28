# accounting/services.py

from datetime import timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook

from .models import (
    LedgerSettings,
    FiscalYear,
    JournalEntry,
    JournalLine,
    Account,
    Journal, Payment, Invoice, PaymentReconciliation,
)

# Common decimal helpers
DECIMAL_ZERO = Decimal("0.000")
QUANTIZER_3DP = Decimal("0.001")


# =====================================================================
# Journal Entry Services
# =====================================================================

def build_lines_from_formset(line_formset):
    """
    Build validated journal lines list from JournalLineFormSet.

    Returns:
        (lines, total_debit, total_credit)
        - lines: list of dicts {account, description, debit, credit, order}
        - total_debit: Decimal
        - total_credit: Decimal

    Raises:
        ValidationError: if there are invalid lines or all lines are empty.
    """
    lines = []
    total_debit = DECIMAL_ZERO
    total_credit = DECIMAL_ZERO

    # Since fields in JournalLineForm are not required, we validate here
    for i, form in enumerate(line_formset.forms):
        if form.cleaned_data.get("DELETE"):
            continue

        account = form.cleaned_data.get("account")
        desc = form.cleaned_data.get("description") or ""

        raw_debit = form.cleaned_data.get("debit")
        raw_credit = form.cleaned_data.get("credit")

        debit = Decimal(str(raw_debit)) if raw_debit is not None else Decimal("0")
        credit = Decimal(str(raw_credit)) if raw_credit is not None else Decimal("0")

        # Quantize to 3 decimal places
        debit = debit.quantize(QUANTIZER_3DP, rounding=ROUND_HALF_UP)
        credit = credit.quantize(QUANTIZER_3DP, rounding=ROUND_HALF_UP)

        # 1) Skip completely empty lines
        if not account and debit == 0 and credit == 0:
            continue

        # 2) Amount without account
        if (debit != 0 or credit != 0) and not account:
            raise ValidationError(
                _("السطر رقم %(row)d: يوجد مبلغ بدون تحديد الحساب.")
                % {"row": i + 1}
            )

        # 3) Account with zero amounts (optional rule – we keep it strict)
        if account and debit == 0 and credit == 0:
            raise ValidationError(
                _("السطر رقم %(row)d: تم تحديد حساب ولكن المبلغ صفر.")
                % {"row": i + 1}
            )

        lines.append(
            {
                "account": account,
                "description": desc,
                "debit": debit,
                "credit": credit,
                "order": i,
            }
        )

        total_debit += debit
        total_credit += credit

    if not lines:
        raise ValidationError(_("يجب إدخال سطر واحد على الأقل."))

    return lines, total_debit, total_credit


# =====================================================================
# Chart of Accounts Services
# =====================================================================

def ensure_default_chart_of_accounts() -> int:
    """
    Create a minimal default chart of accounts if no accounts exist.

    Returns:
        int: number of created accounts (0 if accounts already exist).
    """
    if Account.objects.exists():
        return 0

    accounts_data = [
        # code, name, type, allow_settlement
        ("1000", _("الصندوق"), Account.Type.ASSET, True),
        ("1010", _("البنك"), Account.Type.ASSET, True),
        ("1100", _("العملاء"), Account.Type.ASSET, True),
        ("1200", _("المخزون"), Account.Type.ASSET, False),
        ("2000", _("الموردون"), Account.Type.LIABILITY, True),
        ("3000", _("رأس المال"), Account.Type.EQUITY, False),
        ("4000", _("مبيعات"), Account.Type.REVENUE, False),
        ("5000", _("مشتريات"), Account.Type.EXPENSE, False),
        ("5100", _("مصروفات عمومية وإدارية"), Account.Type.EXPENSE, False),
    ]

    created_count = 0
    for code, name, acc_type, allow_settlement in accounts_data:
        Account.objects.create(
            code=code,
            name=name,
            type=acc_type,
            allow_settlement=allow_settlement,
        )
        created_count += 1

    return created_count


def _parse_bool(value):
    """
    Helper to parse boolean-like values from Excel.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value

    s = str(value).strip().lower()
    if not s:
        return None

    return s in {"1", "true", "yes", "y", "نعم", "صح"}


def import_chart_of_accounts_from_excel(
    file_obj,
    *,
    replace_existing: bool = False,
    fiscal_year: FiscalYear | None = None,
):
    """
    Import chart of accounts from an Excel (.xlsx) file.

    Expected columns (required):
      - code
      - name
      - type  (asset, liability, equity, revenue, expense)

    Optional columns:
      - parent_code
      - allow_settlement
      - is_active
      - opening_debit
      - opening_credit

    If opening balances exist and 'fiscal_year' is provided:
      - a single opening JournalEntry is created with balanced lines.

    Returns:
        dict: {"created": int, "updated": int, "deactivated": int, "errors": list[str]}
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    # --- header mapping ---
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue
        name = str(raw).strip().lower()
        if not name:
            continue
        header_map[name] = col

    required = ["code", "name", "type"]
    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValidationError(
            _("Missing required columns in Excel: %(cols)s")
            % {"cols": ", ".join(missing)}
        )

    rows: list[dict] = []
    seen_codes: set[str] = set()
    errors: list[str] = []

    type_map = {
        "asset": Account.Type.ASSET,
        "liability": Account.Type.LIABILITY,
        "equity": Account.Type.EQUITY,
        "revenue": Account.Type.REVENUE,
        "expense": Account.Type.EXPENSE,
    }

    # --- read rows ---
    for row_idx in range(2, ws.max_row + 1):
        code_val = ws.cell(row=row_idx, column=header_map["code"]).value
        if code_val is None or str(code_val).strip() == "":
            continue

        code = str(code_val).strip()
        name = str(
            ws.cell(row=row_idx, column=header_map["name"]).value or ""
        ).strip()
        type_raw = ws.cell(row=row_idx, column=header_map["type"]).value

        if not name or not type_raw:
            continue

        type_str = str(type_raw).strip().lower()
        if type_str not in type_map:
            errors.append(f"Row {row_idx}: invalid type '{type_str}'")
            continue

        parent_code = None
        if "parent_code" in header_map:
            parent_cell = ws.cell(
                row=row_idx,
                column=header_map["parent_code"],
            ).value
            if parent_cell is not None:
                parent_code = str(parent_cell).strip() or None

        allow_settlement = None
        if "allow_settlement" in header_map:
            val = ws.cell(
                row=row_idx,
                column=header_map["allow_settlement"],
            ).value
            allow_settlement = _parse_bool(val)

        is_active = None
        if "is_active" in header_map:
            val = ws.cell(
                row=row_idx,
                column=header_map["is_active"],
            ).value
            is_active = _parse_bool(val)

        opening_debit = DECIMAL_ZERO
        opening_credit = DECIMAL_ZERO
        if "opening_debit" in header_map:
            val = ws.cell(
                row=row_idx,
                column=header_map["opening_debit"],
            ).value
            if val:
                opening_debit = Decimal(str(val))
        if "opening_credit" in header_map:
            val = ws.cell(
                row=row_idx,
                column=header_map["opening_credit"],
            ).value
            if val:
                opening_credit = Decimal(str(val))

        rows.append(
            {
                "row_idx": row_idx,
                "code": code,
                "name": name,
                "type": type_map[type_str],
                "parent_code": parent_code,
                "allow_settlement": allow_settlement,
                "is_active": is_active,
                "opening_debit": opening_debit,
                "opening_credit": opening_credit,
            }
        )
        seen_codes.add(code)

    created = 0
    updated = 0
    deactivated = 0

    with transaction.atomic():
        accounts_by_code = {a.code: a for a in Account.objects.all()}

        # 1) Create / Update accounts
        for row in rows:
            code = row["code"]
            defaults = {
                "name": row["name"],
                "type": row["type"],
            }
            if row["allow_settlement"] is not None:
                defaults["allow_settlement"] = row["allow_settlement"]
            if row["is_active"] is not None:
                defaults["is_active"] = row["is_active"]

            if code in accounts_by_code:
                acc = accounts_by_code[code]
                for k, v in defaults.items():
                    setattr(acc, k, v)
                # reset parent temporarily – will be set in next loop
                acc.parent = None
                acc.save()
                updated += 1
            else:
                acc = Account.objects.create(code=code, parent=None, **defaults)
                accounts_by_code[code] = acc
                created += 1

        # 2) Set parents
        for row in rows:
            if row["parent_code"]:
                acc = accounts_by_code.get(row["code"])
                parent = accounts_by_code.get(row["parent_code"])
                if acc and parent and acc != parent:
                    acc.parent = parent
                    acc.save(update_fields=["parent"])

        # 3) Deactivate accounts not present in import
        if replace_existing:
            deactivated = Account.objects.exclude(code__in=seen_codes).update(
                is_active=False
            )

        # 4) Opening balances
        opening_lines: list[tuple[Account, Decimal, Decimal]] = []
        tot_dr = DECIMAL_ZERO
        tot_cr = DECIMAL_ZERO

        for row in rows:
            if row["opening_debit"] > 0 or row["opening_credit"] > 0:
                acc = accounts_by_code[row["code"]]
                opening_lines.append(
                    (acc, row["opening_debit"], row["opening_credit"])
                )
                tot_dr += row["opening_debit"]
                tot_cr += row["opening_credit"]

        if opening_lines:
            if not fiscal_year:
                raise ValidationError(
                    _("يجب تحديد سنة مالية لاستيراد الأرصدة الافتتاحية.")
                )

            if tot_dr != tot_cr:
                raise ValidationError(_("القيد الافتتاحي غير متوازن."))

            settings_obj = LedgerSettings.get_solo()
            journal = (
                settings_obj.opening_balance_journal
                or Journal.objects.get_default_for_manual_entry()
            )

            ref = f"OPENING-{fiscal_year.year}"
            # Remove previous imported opening entry for this year/journal
            JournalEntry.objects.filter(journal=journal, reference=ref).delete()

            # Entry date: one day before fiscal year start
            op_date = fiscal_year.start_date - timedelta(days=1)

            entry = JournalEntry.objects.create(
                fiscal_year=None,  # outside current fiscal year
                journal=journal,
                date=op_date,
                reference=ref,
                description=_("رصيد افتتاحي مستورد"),
                posted=True,
                posted_at=timezone.now(),
            )

            for idx, (acc, dr, cr) in enumerate(opening_lines, 1):
                JournalLine.objects.create(
                    entry=entry,
                    account=acc,
                    debit=dr,
                    credit=cr,
                    order=idx,
                )

    return {
        "created": created,
        "updated": updated,
        "deactivated": deactivated,
        "errors": errors,
    }


# =====================================================================
# Sales Invoice Posting Logic
# =====================================================================

def post_sales_invoice_to_ledger(invoice, *, user=None):
    """
    Post a sales invoice to the ledger (JournalEntry + JournalLines).

    - Debit: AR (receivables)
    - Credit: Revenue

    Uses:
      - LedgerSettings.sales_journal
      - LedgerSettings.sales_receivable_account
      - LedgerSettings.sales_revenue_0_account
    """
    if invoice.ledger_entry_id:
        return invoice.ledger_entry

    if invoice.total_amount <= 0:
        raise ValueError(_("لا يمكن ترحيل فاتورة إجماليها صفر."))

    settings = LedgerSettings.get_solo()
    journal = settings.sales_journal
    ar_account = settings.sales_receivable_account
    rev_account = settings.sales_revenue_0_account

    if not all([journal, ar_account, rev_account]):
        raise ValueError(
            _("يرجى ضبط إعدادات دفتر المبيعات وحسابات العملاء والمبيعات.")
        )

    fiscal_year = FiscalYear.for_date(invoice.issued_at)
    if not fiscal_year:
        raise ValueError(_("لا توجد سنة مالية لهذه الفاتورة."))

    with transaction.atomic():
        entry = JournalEntry.objects.create(
            fiscal_year=fiscal_year,
            journal=journal,
            date=invoice.issued_at,
            reference=invoice.display_number,
            description=f"Sales Invoice: {invoice.display_number} - {invoice.customer.name}",
            posted=True,
            posted_at=timezone.now(),
            posted_by=user,
        )

        # Debit AR
        JournalLine.objects.create(
            entry=entry,
            account=ar_account,
            debit=invoice.total_amount,
            credit=0,
            description=f"Inv {invoice.display_number} - Receivable",
        )
        # Credit Revenue
        JournalLine.objects.create(
            entry=entry,
            account=rev_account,
            debit=0,
            credit=invoice.total_amount,
            description=f"Inv {invoice.display_number} - Revenue",
        )

        invoice.ledger_entry = entry
        invoice.save(update_fields=["ledger_entry"])

        return entry


def unpost_sales_invoice_from_ledger(invoice, reversal_date=None, user=None):
    """
    Reverse a posted sales invoice by creating a reversal journal entry.

    Rules:
      - If invoice has any reconcile (paid_amount > 0), reversal is not allowed.
      - Reversal is created on `reversal_date` (or today) in the fiscal year matching that date.
      - Original invoice status is reset back to DRAFT and ledger_entry is cleared.
    """
    if not invoice.ledger_entry:
        return None

    # If invoice has reconcile, do not allow unposting
    if invoice.paid_amount > 0:
        raise ValueError(_("لا يمكن إلغاء ترحيل فاتورة مرتبطة بدفعات."))

    original = invoice.ledger_entry
    rev_date = reversal_date or timezone.now().date()
    fiscal_year = FiscalYear.for_date(rev_date)

    if not fiscal_year:
        raise ValueError(_("لا توجد سنة مالية لتاريخ الإلغاء."))

    with transaction.atomic():
        rev_entry = JournalEntry.objects.create(
            fiscal_year=fiscal_year,
            journal=original.journal,
            date=rev_date,
            reference=f"REV-{original.reference}",
            description=f"Reversal of {original.display_number}",
            posted=True,
            posted_at=timezone.now(),
            posted_by=user,
        )

        for line in original.lines.all():
            JournalLine.objects.create(
                entry=rev_entry,
                account=line.account,
                debit=line.credit,  # swap
                credit=line.debit,  # swap
                description=f"Reversal of line {line.pk}",
            )

        invoice.ledger_entry = None
        invoice.status = invoice.Status.DRAFT
        invoice.save(update_fields=["ledger_entry", "status"])

        return rev_entry


# =====================================================================
# Orders → Invoices
# =====================================================================

def convert_order_to_invoice(order, *, issued_at=None):
    """
    Convert a sales order to an accounting invoice.

    Assumptions about `order`:
      - has `customer`
      - has `items` with (product, description, quantity, unit_price)
      - may have `notes`
      - optionally has fields:
          - invoice (FK to Invoice)
          - STATUS_CONFIRMED (status value for confirmed order)
    """
    InvoiceModel = order._meta.apps.get_model("accounting", "Invoice")
    InvoiceItemModel = order._meta.apps.get_model("accounting", "InvoiceItem")

    with transaction.atomic():
        invoice = InvoiceModel.objects.create(
            customer=order.customer,
            status=InvoiceModel.Status.DRAFT,
            description=order.notes or "",
            issued_at=issued_at or timezone.now(),
        )

        items_to_create = []
        for item in order.items.all():
            items_to_create.append(
                InvoiceItemModel(
                    invoice=invoice,
                    product=item.product,
                    description=item.description,
                    quantity=item.quantity,
                    unit_price=item.unit_price,
                )
            )

        InvoiceItemModel.objects.bulk_create(items_to_create)

        # Recalculate invoice total from items
        invoice.recalculate_totals()

        # Link order → invoice
        order.invoice = invoice

        # Update order status (if a constant is defined on the model)
        status_confirmed = getattr(order, "STATUS_CONFIRMED", None)
        if status_confirmed:
            order.status = status_confirmed

        order.save()

    return invoice


# =====================================================================
# Payment Allocation
# =====================================================================
@transaction.atomic
def allocate_payment_to_invoices(payment: Payment, allocations: Dict[int, Decimal]) -> None:
    """
    تسوية دفعة واحدة مع مجموعة من الفواتير.

    allocations:
        قاموس بشكل:
        {
            invoice_id: amount,
            invoice_id: amount,
            ...
        }

    ملاحظات:
    - لا يتم إنشاء Payment جديد نهائياً.
    - فقط يتم إنشاء/تحديث/حذف أسطر PaymentReconciliation.
    """

    # 1) التحقق من المجموع الكلي للمبالغ
    total_alloc = sum(allocations.values())
    if total_alloc > payment.amount:
        raise ValidationError(_("إجمالي المبالغ المخصصة أكبر من مبلغ الدفعة."))

    # 2) التحقق من كل فاتورة
    for invoice_id, amount in allocations.items():
        if amount <= 0:
            # بنسمح بقيم <= صفر كإشارة للحذف لاحقاً
            continue

        try:
            invoice = Invoice.objects.get(pk=invoice_id)
        except Invoice.DoesNotExist:
            raise ValidationError(_("الفاتورة برقم %(inv)s غير موجودة.") % {"inv": invoice_id})

        # الطرف لازم يكون نفسه
        if invoice.customer != payment.contact:
            raise ValidationError(
                _("الفاتورة %(inv)s ليست لنفس الطرف المرتبط بالدفعة.")
                % {"inv": invoice.display_number}
            )

        # لا يتجاوز الرصيد المتبقي على الفاتورة
        if amount > invoice.balance:
            raise ValidationError(
                _("المبلغ المخصص للفاتورة %(inv)s أكبر من رصيدها المتبقي.")
                % {"inv": invoice.display_number}
            )

        # لا يتجاوز المبلغ غير المخصص من الدفعة (تحقق إضافي)
        if amount > payment.unallocated_amount:
            raise ValidationError(
                _("المبلغ المخصص أكبر من المبلغ غير المخصص المتبقي في الدفعة.")
            )

    # 3) إنشاء/تحديث/حذف التسويات
    for invoice_id, amount in allocations.items():
        invoice = Invoice.objects.get(pk=invoice_id)

        if amount <= 0:
            # لو صفر أو أقل: نحذف أي تسوية موجودة لهذه الفاتورة مع هذه الدفعة
            PaymentReconciliation.objects.filter(payment=payment, invoice=invoice).delete()
            continue

        # update_or_create: يعدّل لو موجود، ينشئ لو جديد
        PaymentReconciliation.objects.update_or_create(
            payment=payment,
            invoice=invoice,
            defaults={"amount": amount},
        )

        # بعد كل تعديل، ممكن تختار تحدث حالة الفاتورة:
        invoice.update_payment_status()


@transaction.atomic
def clear_payment_allocations(payment: Payment) -> None:
    """
    إلغاء جميع تسويات الدفعة (تحريرها من كل الفواتير).
    لا يحذف الدفعة نفسها.
    """
    qs = PaymentReconciliation.objects.filter(payment=payment)

    # نحفظ قائمة الفواتير لتحديث حالتها بعد الحذف
    invoices = list({alloc.invoice for alloc in qs})

    qs.delete()

    for invoice in invoices:
        invoice.update_payment_status()
